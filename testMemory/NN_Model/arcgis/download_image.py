#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2018-09-07 16:38:49
# @Author  : ${menzec} (${menzc@outlook.com})
# @Link    : http://example.org
# @Version : $Id$

import os
from openpyxl import Workbook, load_workbook


def find_all_index(alist, vaule, times):
    index = []
    left_index = 0
    right_index = len(alist)
    for i in range(times):
        left_index = alist.index(vaule, left_index, right_index)
        index.append(left_index)
        left_index += 1
    return index


def make_list(datafile):
    data_fn = load_workbook(datafile)
    sheet_data = data_fn['Sheet1']
    coordinate = data_fn.create_sheet('coordinate')
    data_sh_row_now = 1
    data_sh_row_last = 1
    data_sh_col = 1
    cor_row = 1
    cor_col = 2
    cor_data_sh_col = 21
    while sheet_data.cell(data_sh_row_now, data_sh_col).value:
        if sheet_data.cell(data_sh_row_now, data_sh_col).value != sheet_data.cell(data_sh_row_now + 1, data_sh_col).value:
            coordinate.cell(cor_row, 1).value = sheet_data.cell(
                data_sh_row_now, data_sh_col).value
            for m, i in enumerate(range(data_sh_row_last, data_sh_row_now + 1)):
                coordinate.cell(
                    cor_row, cor_col + m).value = double(sheet_data.cell(i, cor_data_sh_col).value)
                coordinate.cell(
                    cor_row, cor_col + m + data_sh_row_now - data_sh_row_last + 1).value = double(sheet_data.cell(i, cor_data_sh_col + 1).value)
                print(('%d %d %f,%d %d %f') % (cor_row, cor_col + m, sheet_data.cell(i, cor_data_sh_col).value,
                                               cor_row, cor_col + m + data_sh_row_now - data_sh_row_last + 1, sheet_data.cell(i, cor_data_sh_col + 1).value))
            cor_row += 1
            data_sh_row_last = data_sh_row_now + 1
        data_sh_row_now += 1
    data_fn.save(datafile)


def main():
    make_list(r'D:\data\test.xlsx')
    # data_fn = load_workbook(r'D:\data\test.xlsx')
    # sheet_data = data_fn['Sheet1']
    # coordinate = data_fn.create_sheet('coordinate')
    # coordinate.cell(2, 2).value = 505
    # point = coordinate.cell(2, 2).value
    # print(point)
    # coordinate.cell(2, 3).value = point
    # data_fn.save(r'D:\data\test.xlsx')

if __name__ == '__main__':
    main()
